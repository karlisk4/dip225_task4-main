import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook, load_workbook 

empl_name = input()
service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(options=option)

name=[]
with open("people.csv", "r") as file:
    next(file)
    for line in file:
        row=line.rstrip().split(",")
        row_name = [row[2], row[3]] 
        name.append(row_name)

url = "https://emn178.github.io/online-tools/crc32.html"
driver.get(url)
time.sleep(1)

wb=load_workbook('salary.xlsx')
ws=wb.active
hashed_sal=[]
rows = ws.max_row+1
for row in range(2,rows):
    hash=ws['A'+str(row)].value
    salary=ws['B'+str(row)].value
    if hash:
        hashed = [hash, salary]
        hashed_sal.append(hashed)

name_hash = []
for el in name:
    find = driver.find_element(By.ID, "input")
    find.clear()
    full_name = str(el[0] + " " + el[1])
    find.send_keys(full_name)
    find = driver.find_element(By.ID, "output")
    temp = find.get_attribute("value")
    temp_merge = [full_name, temp]
    name_hash.append(temp_merge)

sal = 0
sal_list = []
for elem in name_hash:
    for a in hashed_sal:
        if elem[1] == a[0]:
            sal = sal + int(a[1])
    sal_val = [elem[0], sal]
    sal_list.append(sal_val)
    sal = 0

for ele in sal_list:
    if ele[0] == empl_name:
        print(str(ele[0]) + " " + str(ele[1]))